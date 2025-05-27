from flask import Flask, render_template, request, redirect, url_for, send_file, session
from datetime import datetime
import openpyxl
import os
import math
import base64
from PIL import Image

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # For session login

# Constants
OFFICE_LAT = 22.957531430064837
OFFICE_LNG = 72.65531222328558
ALLOWED_DISTANCE = 100  # meters
EMPLOYEES = ['SUJAL', 'CA SHENAL', 'CA HEENA', 'SHEFALI', 'SAPNA', 'SONALI']
ADMIN_NAME = 'CA SHENAL'
EXCEL_FILE = 'attendance.xlsx'

# Ensure Excel file exists
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Name', 'Date', 'Time', 'Type', 'Latitude', 'Longitude', 'Image'])
    wb.save(EXCEL_FILE)

# Haversine formula to calculate distance
def calculate_distance(lat1, lon1, lat2, lon2):
    R = 6371000  # Radius of the Earth in meters
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = math.sin(delta_phi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(delta_lambda/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c

# Employee route
@app.route('/attendance/<name>', methods=['GET'])
def attendance(name):
    name_upper = name.upper()
    if name_upper not in EMPLOYEES:
        return "❌ Unauthorized User", 403
    return render_template("employee.html", name=name_upper)

# Punch route
@app.route('/punch', methods=['POST'])
def punch():
    name = request.form['name'].upper()
    lat = float(request.form['lat'])
    lng = float(request.form['lng'])
    img_data = request.form['image']
    punch_type = request.form['type']
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    distance = calculate_distance(lat, lng, OFFICE_LAT, OFFICE_LNG)
    if distance > ALLOWED_DISTANCE:
        return "❌ You are not within 100 meters of the office.", 400

    # Decode and save photo
    os.makedirs('photos', exist_ok=True)
    img_bytes = base64.b64decode(img_data.split(',')[1])
    img_filename = f"photos/{name}_{date_str}_{punch_type}.png"
    with open(img_filename, 'wb') as f:
        f.write(img_bytes)

    # Load or create Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Check if already punched
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == name and row[1] == date_str and row[3] == punch_type:
            return f"⚠️ Already punched {punch_type} today.", 400

    # Save to Excel
    ws.append([name, date_str, time_str, punch_type, lat, lng, img_filename])
    wb.save(EXCEL_FILE)

    return redirect(url_for('attendance', name=name))

# Admin login
@app.route('/admin', methods=['GET', 'POST'])
def admin():
    if request.method == 'POST':
        if request.form['name'].upper() == ADMIN_NAME:
            session['admin'] = True
            return redirect(url_for('dashboard'))
        else:
            return "❌ Access Denied", 403
    return render_template("admin_login.html")

# Admin dashboard
@app.route('/dashboard')
def dashboard():
    if not session.get('admin'):
        return redirect(url_for('admin'))
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    records = [row for row in ws.iter_rows(min_row=2, values_only=True)]
    return render_template("dashboard.html", records=records)

# Download Excel
@app.route('/download')
def download():
    if not session.get('admin'):
        return redirect(url_for('admin'))
    return send_file(EXCEL_FILE, as_attachment=True)

# Run app
if __name__ == "__main__":
    app.run(debug=True)
